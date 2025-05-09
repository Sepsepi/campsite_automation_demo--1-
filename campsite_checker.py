import time
import openpyxl # Changed from pandas
from playwright.sync_api import sync_playwright
from datetime import date, timedelta, datetime
import tkinter as tk
from tkinter import filedialog # Added for save dialog
import sys
import os
# gui.py will import perform_site_check from this file
# from gui import CampsiteCheckerGUI # This line will be moved to main() to avoid circular import at module level

# --- Configuration ---
URL = "https://nysuffolkctyweb.myvscloud.com/webtrac/web/search.html?Module=RN&Type=FamilySite&webscreendesign=webtrac%20screen%20design&Display=Detail"
# TARGET_PARK_VALUE and TARGET_SITES are set from the GUI, defaulting only at app launch
OUTPUT_FILE = "campsite_availability.xlsx"
DATE_FORMAT = "%m/%d/%Y" # Format expected by the website input
REQUEST_DELAY_SECONDS = 1 # Delay between checking each date

# --- Selectors ---
PARK_LABEL_SELECTOR = "#category_vm_2_label"
PARK_LIST_ITEM_SELECTOR = "#category_vm_2_wrap li[data-value='{}']"  # Use .format(park_name) at runtime
SITE_LIST_ITEM_SELECTOR = "#keyword_vm_3_wrap li[data-value='{}']"  # Use .format(site_number) at runtime
NIGHTS_LABEL_SELECTOR = "#nights_vm_1_label"
NIGHTS_BUTTON_SELECTOR = "#nights_vm_1_button"
NIGHTS_OPTION_1_SELECTOR = "#nights_vm_1_wrap li[data-value='1']"
NIGHTS_OPTION_4_SELECTOR = "#nights_vm_1_wrap li[data-value='4']"
KEYWORD_LABEL_SELECTOR = "label[for='keyword']"
KEYWORD_INPUT_SELECTOR = "#keyword"
BEGIN_DATE_LABEL_SELECTOR = "#begindate_vm_6_label"
BEGIN_DATE_INPUT_SELECTOR = "#begindate" # Hidden input
SEARCH_BUTTON_SELECTOR = "#rnwebsearch_buttonsearch"
RESULTS_CONTAINER_SELECTOR = "div.result-content" # To wait for results loading

# --- JavaScript to Extract Data ---
# EXTRACT_JS now uses a placeholder for site_list, which is replaced at runtime
EXTRACT_JS = """
(() => {
    const results = [];
    const resultBlocks = document.querySelectorAll('div.result-content');
    const targetSiteNumbers = SITE_LIST_PLACEHOLDER.split(',').map(s => s.trim());

    resultBlocks.forEach(block => {
        const siteNameElement = block.querySelector('.result-header__info h2 span');
        const siteName = siteNameElement ? siteNameElement.textContent.trim() : null;
        const siteNumberMatch = siteName ? siteName.match(/Site (\\d+)/) : null;
        const siteNumber = siteNumberMatch ? siteNumberMatch[1] : null;

        if (siteNumber) {
            const addButton = block.querySelector('td.button-cell--cart a.button');
            let status = 'Unknown';
            if (addButton) {
                if (addButton.classList.contains('success')) {
                    status = 'Available';
                } else if (addButton.classList.contains('error')) {
                    status = 'Unavailable';
                }
            }
            if (targetSiteNumbers.includes(siteNumber)) {
                 results.push({ site: siteNumber, status: status });
            }
        }
    });
    return results;
})();
"""

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def perform_site_check(start_date_str, end_date_str, log_callback, progress_callback=None, headless=False, park_name=None, site_list=None): # Add headless, park_name, and site_list parameters
    # Use the value from the GUI box directly for park_name and site_list, defaulting only if not provided
    park_name = park_name if park_name else "Smith Point"
    site_list = site_list if site_list else "230, 234, 236, 238, 240, 242, 244, 246, 248, 250, 252, 254, 256, 258, 260, 262, 264, 266, 268, 270"
    log_callback("Starting site check from {} to {} (headless={}, park={}, sites={})...".format(start_date_str, end_date_str, headless, park_name, site_list))
    all_results = []
    
    try:
        start_date_obj = datetime.strptime(start_date_str, DATE_FORMAT).date()
        end_date_obj = datetime.strptime(end_date_str, DATE_FORMAT).date()
    except ValueError:
        log_callback("Error: Invalid date format. Please use {}.".format(DATE_FORMAT.replace('%m','MM').replace('%d','DD').replace('%Y','YYYY')))
        return

    current_date = start_date_obj
    today = date.today()

    total_days = (end_date_obj - start_date_obj).days + 1
    processed_days = 0

    if progress_callback:
        progress_callback(0) # Initial progress

    with sync_playwright() as p:
        try:
            # For development, ensure Playwright browsers are installed: playwright install
            # For cx_Freeze, if bundling, executablePath might be needed.
            # For now, assume browsers are available in the environment.
            browser = p.chromium.launch(headless=headless) 
            page = browser.new_page()
            log_callback("Navigating to {}...".format(URL))
            page.goto(URL, wait_until='networkidle')
            log_callback("Page loaded.")

            log_callback("Performing initial setup clicks...")
            try:
                page.locator(PARK_LABEL_SELECTOR).click()
                park_name_clean = park_name.strip()
                # Try to select by data-value attribute first
                park_selector = PARK_LIST_ITEM_SELECTOR.format(park_name_clean)
                park_elem = page.locator(park_selector)
                if park_elem.count() > 0:
                    park_elem.click()
                    log_callback(f"- Selected Park by data-value: {park_name_clean}")
                else:
                    # Fallback: try to match visible text (case-insensitive)
                    all_parks = page.locator("#category_vm_2_wrap li").all()
                    matched = False
                    for park_option in all_parks:
                        value = park_option.get_attribute('data-value')
                        text = park_option.inner_text().strip() if hasattr(park_option, 'inner_text') else None
                        if text and text.lower() == park_name_clean.lower():
                            park_option.click()
                            log_callback(f"- Selected Park by visible text: {text}")
                            matched = True
                            break
                        elif value and value.strip().lower() == park_name_clean.lower():
                            park_option.click()
                            log_callback(f"- Selected Park by data-value fallback: {value}")
                            matched = True
                            break
                    if not matched:
                        error_msg = f"Park '{park_name_clean}' not found in the park list on the website. Please select a valid park name as shown on the website."
                        log_callback(f"ERROR: {error_msg}")
                        print(error_msg, file=sys.stderr)
                        return
            except Exception as e:
                error_msg = f"Could not select park '{park_name}': {e}"
                log_callback(f"ERROR: {error_msg}")
                print(error_msg, file=sys.stderr)
                return
            try:
                page.locator(KEYWORD_LABEL_SELECTOR).click()
                site_list_clean = site_list.strip() if site_list else "230, 234, 236, 238, 240, 242, 244, 246, 248, 250, 252, 254, 256, 258, 260, 262, 264, 266, 268, 270"
                # Remove duplicate site numbers and keep order
                site_numbers = [s.strip() for s in site_list_clean.split(',') if s.strip().isdigit()]
                seen = set()
                unique_sites = []
                for num in site_numbers:
                    if num not in seen:
                        unique_sites.append(num)
                        seen.add(num)
                # Try to select each site in the dropdown if it exists
                for site_num in unique_sites:
                    site_selector = SITE_LIST_ITEM_SELECTOR.format(site_num)
                    site_elem = page.locator(site_selector)
                    if site_elem.count() > 0:
                        site_elem.click()
                        log_callback(f"- Selected Site in dropdown: {site_num}")
                # Always fill the input as well (for keyword search)
                # Sort site numbers numerically and join as string for input
                try:
                    sorted_sites = sorted(unique_sites, key=lambda x: int(x))
                except Exception:
                    sorted_sites = unique_sites  # fallback to original order if conversion fails
                final_site_list = ', '.join(sorted_sites)
                page.locator(KEYWORD_INPUT_SELECTOR).fill(final_site_list)
                log_callback(f"- Entered Site Keywords: {final_site_list}")
                # Update the site_list variable for JS extraction to match what was actually searched
                site_list = final_site_list
            except Exception as e:
                error_msg = f"Could not enter site list '{site_list}': {e}"
                log_callback(f"ERROR: {error_msg}")
                print(error_msg, file=sys.stderr)
                return
            log_callback("Initial setup complete.")

            while current_date <= end_date_obj:
                loop_date_str = current_date.strftime(DATE_FORMAT)
                log_callback("Checking date: {}...".format(loop_date_str))

                try:
                    days_diff = (current_date - today).days
                    nights_to_select = 4 if days_diff > 30 else 1
                    log_callback("- Setting Nights to {} (Date is {} days from today)".format(nights_to_select, days_diff))

                    page.locator(NIGHTS_LABEL_SELECTOR).click()
                    page.locator(NIGHTS_BUTTON_SELECTOR).click() # Assuming this reveals the options
                    if nights_to_select == 4:
                        page.locator(NIGHTS_OPTION_4_SELECTOR).click()
                    else:
                        page.locator(NIGHTS_OPTION_1_SELECTOR).click()

                    page.locator(BEGIN_DATE_LABEL_SELECTOR).click()
                    page.evaluate("""
                        const dateInput = document.querySelector('{}');
                        if (dateInput) {{
                            dateInput.value = '{}';
                            dateInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
                        }} else {{
                            console.error('Date input not found for {}!');
                        }}
                    """.format(BEGIN_DATE_INPUT_SELECTOR, loop_date_str, loop_date_str))
                    log_callback("- Set Begin Date to {} via JS".format(loop_date_str))

                    page.locator(SEARCH_BUTTON_SELECTOR).click()
                    log_callback("- Clicked Search")

                    page.wait_for_load_state('networkidle', timeout=15000)
                    log_callback("- Results loaded (network idle)")

                    # Replace SITE_LIST_PLACEHOLDER with the actual site_list for this run
                    js_to_run = EXTRACT_JS.replace('SITE_LIST_PLACEHOLDER', f"'{site_list}'")
                    daily_results = page.evaluate(js_to_run)
                    log_callback("- Extracted {} results for {}".format(len(daily_results), loop_date_str))

                    for result in daily_results:
                        result['date'] = loop_date_str
                        result['nights_searched'] = nights_to_select
                        all_results.append(result)

                except Exception as e:
                    log_callback("!! Error processing date {}: {}".format(loop_date_str, e))

                current_date += timedelta(days=1)
                processed_days += 1
                if progress_callback:
                    progress_percentage = (processed_days / total_days) * 100
                    progress_callback(progress_percentage)

                if current_date <= end_date_obj: # Only sleep if there are more dates
                    log_callback("Waiting {}s before next request...".format(REQUEST_DELAY_SECONDS))
                    time.sleep(REQUEST_DELAY_SECONDS)
            
            log_callback("All dates processed.")
            if progress_callback:
                progress_callback(100) # Ensure it hits 100% at the end of loop

        except Exception as e:
            log_callback("An error occurred during the Playwright process: {}".format(e))
        finally:
            if 'browser' in locals() and browser.is_connected():
                log_callback("Closing browser...")
                browser.close()

    if all_results:
        log_callback("\nCollected {} total availability records.".format(len(all_results)))
        log_callback("Processing data and saving to Excel using openpyxl...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Availability"
        
        # Write header
        header = ['date', 'site', 'status', 'nights_searched']
        ws.append(header)
        
        # Write data rows
        for result_item in all_results:
            # Ensure order matches header
            row_data = [
                result_item.get('date'),
                result_item.get('site'),
                result_item.get('status'),
                result_item.get('nights_searched')
            ]
            ws.append(row_data)
        
        # --- IMPROVED FORMATTING ---
        from openpyxl.styles import Font, Alignment, PatternFill
        # Header style
        header_font = Font(bold=True, size=13)
        header_fill = PatternFill("solid", fgColor="D9E1F2")  # Light blue fill
        header_alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        # No borders for data rows or header
        # Adjust column widths (make them wider for more space)
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 8  # Add more padding for width
        # Freeze header row
        ws.freeze_panes = ws["A2"]
        # --- END IMPROVED FORMATTING ---

        # Ask user for save location
        # Reformat dates for filename to avoid slashes
        try:
            s_date_obj = datetime.strptime(start_date_str, DATE_FORMAT)
            e_date_obj = datetime.strptime(end_date_str, DATE_FORMAT)
            formatted_start_date = s_date_obj.strftime("%m-%d-%Y")
            formatted_end_date = e_date_obj.strftime("%m-%d-%Y")
            default_filename = "Campsite_Availability_{}_to_{}.xlsx".format(formatted_start_date, formatted_end_date)
        except ValueError: # Fallback if date parsing fails for some reason
            default_filename = "Campsite_Availability.xlsx"

        # --- HISTORY FOLDER FEATURE ---
        base_dir = os.path.dirname(os.path.abspath(__file__))
        history_dir = os.path.join(base_dir, "History")
        if not os.path.exists(history_dir):
            os.makedirs(history_dir)
        # --- END HISTORY FOLDER FEATURE ---

        output_path = filedialog.asksaveasfilename(
            initialdir=history_dir, # Set the default directory to History
            initialfile=default_filename, # Set the default filename
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Campsite Availability As..."
        )

        if output_path: # If the user selected a path (didn't cancel)
            try:
                wb.save(output_path)
                log_callback("Data successfully saved to {}".format(output_path))
                # --- HISTORY FEATURE ---
                # Create History folder if it doesn't exist (next to APP)
                base_dir = os.path.dirname(os.path.abspath(__file__))
                history_dir = os.path.join(base_dir, "History")
                if not os.path.exists(history_dir):
                    os.makedirs(history_dir)
                # Write a log file for this session
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                log_filename = "history_{}_to_{}_{}.txt".format(formatted_start_date, formatted_end_date, timestamp)
                log_path = os.path.join(history_dir, log_filename)
                with open(log_path, "w", encoding="utf-8") as f:
                    f.write("Scrape session: {}\n".format(timestamp))
                    f.write("Start date: {}\n".format(start_date_str))
                    f.write("End date: {}\n".format(end_date_str))
                    f.write("Excel file: {}\n".format(output_path))
                    f.write("Records collected: {}\n".format(len(all_results)))
                log_callback("History log saved to {}".format(log_path))
                # --- END HISTORY FEATURE ---
            except Exception as e:
                log_callback("Error saving Excel file to {}: {}".format(output_path, e))
        else:
            log_callback("Save operation cancelled by user.")
            
    else:
        log_callback("No data collected. Check logs for any issues.")
    
    log_callback("Site check script finished.")
    # Ensure progress is 100 if everything completes, even if no results
    if progress_callback:
        progress_callback(100)


def main():
    # Import gui here to avoid circular import at module level
    from gui import CampsiteCheckerGUI 
    app = CampsiteCheckerGUI()
    app.mainloop()

if __name__ == "__main__":
    # This is the main entry point when running the script directly
    # For cx_Freeze, it will also be the entry point of the executable
    main()